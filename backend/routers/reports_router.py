"""Activity log + reports (PDF + Excel) routes."""
import io
from datetime import datetime
from typing import Optional
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from deps import get_db, require_super_admin

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.units import mm
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

router = APIRouter(prefix="/activity", tags=["activity"])


@router.get("")
async def list_activity(user: dict = Depends(require_super_admin),
                        limit: int = 200,
                        target_collection: Optional[str] = None,
                        admin_id: Optional[str] = None):
    db = get_db()
    q = {}
    if target_collection:
        q["target_collection"] = target_collection
    if admin_id:
        q["admin_id"] = admin_id
    docs = await db.activity_logs.find(q, {"_id": 0}).sort("created_at", -1).to_list(limit)
    return docs


# ---------- Reports ----------
reports_router = APIRouter(prefix="/reports", tags=["reports"])


def _fmt_inr(n: float) -> str:
    n = float(n or 0)
    s = f"{n:,.0f}"
    return f"Rs {s}"


async def _gather_report(db, month: Optional[str] = None):
    """
    Gather complete booking records and payables ledger.
    If month is 'all' or empty, returns all historical bookings.
    """
    is_all_time = not month or month.strip().lower() in ("all", "all-time", "")
    if is_all_time:
        q = {}
        period_label = "Complete History (All Time)"
        file_label = "all-time"
    else:
        q = {"start_date": {"$regex": f"^{month}"}}
        period_label = f"Month: {month}"
        file_label = month

    bookings = await db.bookings.find(q, {"_id": 0}).sort("start_date", 1).to_list(10000)
    settings = await db.settings.find_one({"id": "default"}) or {}
    savings_pct = float(settings.get("savings_percent", 10))

    owners = await db.car_owners.find({}, {"_id": 0}).to_list(500)
    agents = await db.agents.find({}, {"_id": 0}).to_list(500)
    car_map = {c["id"]: c for c in await db.cars.find({}, {"_id": 0}).to_list(500)}
    owner_map = {o["id"]: o for o in owners}
    agent_map = {a["id"]: a for a in agents}

    total_income = 0.0
    total_owner_cost = 0.0
    total_agent_fee = 0.0
    total_margin = 0.0
    total_net = 0.0
    total_cash_income = 0.0
    total_online_income = 0.0
    total_deposit_held = 0.0
    total_deposit_refunded = 0.0

    for b in bookings:
        car = car_map.get(b.get("car_id"), {}) if b.get("car_id") else {}
        owner = owner_map.get(b.get("owner_id"), {}) if b.get("owner_id") else {}
        agent = agent_map.get(b.get("assigned_agent_id"), {}) if b.get("assigned_agent_id") else {}

        # Resolve car model and plate accurately
        if not b.get("car_model") or b.get("car_model") == "—":
            b["car_model"] = car.get("model", "Standard Vehicle")
        if not b.get("car_registration") or b.get("car_registration") == "—":
            b["car_registration"] = car.get("registration_no", "TBD")
        if not b.get("owner_name") or b.get("owner_name") == "—":
            b["owner_name"] = car.get("owner_name") or owner.get("name", "—")
        if not b.get("agent_name") or b.get("agent_name") == "—":
            b["agent_name"] = agent.get("name", "—")

        c_rate = float(b.get("customer_rate") or 0)
        o_cost = float(b.get("cost_rate") or 0)
        a_fee = float(b.get("agent_fee") or 0)
        marg = float(b.get("margin") if b.get("margin") is not None else (c_rate - o_cost))
        net = float(b.get("net_profit") if b.get("net_profit") is not None else (marg - a_fee))
        dep_amt = float(b.get("deposit_amount") or 0)
        dep_st = b.get("deposit_status") or "none"
        pay_m = b.get("payment_method") or "cash"

        b["customer_rate"] = c_rate
        b["cost_rate"] = o_cost
        b["agent_fee"] = a_fee
        b["margin"] = marg
        b["net_profit"] = net
        b["deposit_amount"] = dep_amt
        b["deposit_status"] = dep_st
        b["payment_method"] = pay_m
        b["days"] = int(b.get("days") or 1)

        total_income += c_rate
        total_owner_cost += o_cost
        total_agent_fee += a_fee
        total_margin += marg
        total_net += net

        if pay_m == "cash":
            total_cash_income += c_rate
        else:
            total_online_income += c_rate

        if dep_st == "received":
            total_deposit_held += dep_amt
        elif dep_st == "refunded":
            total_deposit_refunded += dep_amt

    savings = total_net * (savings_pct / 100.0)

    return {
        "month": month or "all",
        "period_label": period_label,
        "file_label": file_label,
        "is_all_time": is_all_time,
        "bookings": bookings,
        "owners": owners,
        "agents": agents,
        "generated_at": datetime.now().strftime("%d %b %Y, %I:%M %p"),
        "totals": {
            "income": total_income,
            "owner_cost": total_owner_cost,
            "agent_fee": total_agent_fee,
            "margin": total_margin,
            "net_profit": total_net,
            "cash_income": total_cash_income,
            "online_income": total_online_income,
            "deposit_held": total_deposit_held,
            "deposit_refunded": total_deposit_refunded,
            "savings": savings,
            "savings_percent": savings_pct,
        },
    }


@reports_router.get("/monthly.pdf")
async def monthly_pdf(month: Optional[str] = Query(None, description="YYYY-MM or 'all'"),
                      user: dict = Depends(require_super_admin)):
    db = get_db()
    data = await _gather_report(db, month)

    buf = io.BytesIO()
    # Landscape A4 gives generous width for all columns
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=12 * mm,
        bottomMargin=10 * mm,
    )
    styles = getSampleStyleSheet()
    slate900 = colors.HexColor("#20373B")
    slate700 = colors.HexColor("#2C494E")
    slate500 = colors.HexColor("#64748B")
    slate100 = colors.HexColor("#F4FAFC")
    gold = colors.HexColor("#D97706")
    emerald = colors.HexColor("#047857")
    red = colors.HexColor("#B91C1C")

    title_style = ParagraphStyle(
        "t", parent=styles["Title"],
        textColor=slate900, fontSize=18, spaceAfter=2, alignment=0, fontName="Helvetica-Bold"
    )
    sub_style = ParagraphStyle(
        "s", parent=styles["Normal"],
        textColor=slate500, fontSize=9, spaceAfter=10
    )
    h2 = ParagraphStyle(
        "h2", parent=styles["Heading2"],
        textColor=slate900, fontSize=12, spaceAfter=6, spaceBefore=12, fontName="Helvetica-Bold"
    )
    brand = ParagraphStyle(
        "brand", parent=styles["Normal"],
        textColor=gold, fontSize=10, spaceAfter=2, fontName="Helvetica-Bold"
    )
    cell_txt = ParagraphStyle(
        "cell_txt", parent=styles["Normal"],
        fontSize=7.5, leading=9, textColor=slate900
    )
    cell_txt_bold = ParagraphStyle(
        "cell_txt_bold", parent=styles["Normal"],
        fontSize=7.5, leading=9, textColor=slate900, fontName="Helvetica-Bold"
    )
    cell_txt_sub = ParagraphStyle(
        "cell_txt_sub", parent=styles["Normal"],
        fontSize=6.5, leading=8, textColor=slate500
    )

    story = []
    story.append(Paragraph("CAR CASTLE GOA — FLEET & FINANCIAL REPORT", brand))
    story.append(Paragraph(f"{data['period_label']}", title_style))
    story.append(Paragraph(f"Generated on {data['generated_at']} · Self-Drive Rentals, Fleet Ledger & Airport Transfers", sub_style))

    t = data["totals"]
    # KPI Grid
    kpi_data = [
        ["Total Bookings", "Gross Sales", "Owner Payables", "Net Margin", "Cash / Online", "Deposits Held"],
        [
            str(len(data["bookings"])),
            _fmt_inr(t["income"]),
            _fmt_inr(t["owner_cost"]),
            _fmt_inr(t["net_profit"]),
            f"Cash: {_fmt_inr(t['cash_income'])}\nOnline: {_fmt_inr(t['online_income'])}",
            f"Held: {_fmt_inr(t['deposit_held'])}\nRef.: {_fmt_inr(t['deposit_refunded'])}",
        ],
    ]
    kpi_table = Table(kpi_data, colWidths=[45 * mm] * 6)
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), slate900),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BACKGROUND", (0, 1), (-1, 1), slate100),
        ("TEXTCOLOR", (3, 1), (3, 1), emerald),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#C3E7F1")),
    ]))
    story.append(kpi_table)

    # Bookings Master Table
    story.append(Paragraph(f"Bookings ({len(data['bookings'])})", h2))
    rows = [
        ["Dates & Times", "Customer", "Car & Plate", "Owner", "Mode", "Deposit", "Owner Rent", "Customer Rate", "Net Profit", "Status"]
    ]
    for b in data["bookings"]:
        dates_p = Paragraph(f"<b>{b.get('start_date','')[:10]}</b> ({b.get('pickup_time','09:00')})<br/>→ <b>{b.get('end_date','')[:10]}</b> ({b.get('drop_time','09:00')})<br/><font color='#519CAB'>{b.get('days',1)}d</font>", cell_txt)
        cust_p = Paragraph(f"<b>{b.get('customer_name','')}</b><br/><font color='#64748B'>{b.get('customer_contact','')}</font>", cell_txt)
        car_p = Paragraph(f"<b>{b.get('car_model','')}</b><br/><font color='#519CAB'><b>{b.get('car_registration','TBD')}</b></font>", cell_txt)
        owner_p = Paragraph(f"{b.get('owner_name','—')}", cell_txt)
        mode_p = Paragraph("Cash" if b.get("payment_method") == "cash" else "Online", cell_txt)
        
        dep_amt = float(b.get("deposit_amount") or 0)
        dep_str = f"Rs {dep_amt:,.0f} ({b.get('deposit_status','none')})" if dep_amt > 0 else "—"
        dep_p = Paragraph(dep_str, cell_txt)

        rows.append([
            dates_p,
            cust_p,
            car_p,
            owner_p,
            mode_p,
            dep_p,
            _fmt_inr(b["cost_rate"]),
            _fmt_inr(b["customer_rate"]),
            _fmt_inr(b["net_profit"]),
            b.get("status", "").replace("_", " ").title(),
        ])

    if len(rows) == 1:
        rows.append(["—", "No bookings found for this period", "", "", "", "", "", "", "", ""])

    col_widths = [40 * mm, 38 * mm, 38 * mm, 30 * mm, 18 * mm, 24 * mm, 23 * mm, 24 * mm, 23 * mm, 20 * mm]
    bt = Table(rows, colWidths=col_widths, repeatRows=1)
    bt.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), slate100),
        ("TEXTCOLOR", (0, 0), (-1, 0), slate900),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("ALIGN", (6, 0), (8, -1), "RIGHT"),
        ("ALIGN", (9, 0), (9, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E2E8F0")),
        ("TEXTCOLOR", (6, 1), (6, -1), red),
        ("TEXTCOLOR", (8, 1), (8, -1), emerald),
    ]))
    story.append(bt)

    # Outstanding Owner Payables
    story.append(Paragraph("Outstanding Owner Payables Ledger", h2))
    o_rows = [["Owner", "Contact", "Total Owed", "Total Paid", "Outstanding Balance"]]
    for o in data["owners"]:
        owed = float(o.get("total_owed") or 0)
        paid = float(o.get("total_paid") or 0)
        bal = owed - paid
        if bal > 0.01 or owed > 0:
            o_rows.append([o.get("name", "—"), o.get("contact", "—"), _fmt_inr(owed), _fmt_inr(paid), _fmt_inr(bal)])
    if len(o_rows) == 1:
        o_rows.append(["—", "All accounts settled", "Rs 0", "Rs 0", "Rs 0"])
    ot = Table(o_rows, colWidths=[65 * mm, 55 * mm, 50 * mm, 50 * mm, 55 * mm], repeatRows=1)
    ot.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), slate100),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E2E8F0")),
        ("TEXTCOLOR", (4, 1), (4, -1), red),
    ]))
    story.append(ot)

    # Footer
    story.append(Spacer(1, 8))
    footer = ParagraphStyle("f", parent=styles["Normal"], textColor=slate500, fontSize=8, alignment=1)
    story.append(Paragraph(f"Car Castle Goa · System Generated Report · {data['period_label']}", footer))

    doc.build(story)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="car-castle-goa-{data["file_label"]}.pdf"'},
    )


@reports_router.get("/monthly.xlsx")
async def monthly_xlsx(month: Optional[str] = Query(None, description="YYYY-MM or 'all'"),
                       user: dict = Depends(require_super_admin)):
    db = get_db()
    data = await _gather_report(db, month)

    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="20373B")
    header_font = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
    sub_fill = PatternFill("solid", fgColor="F4FAFC")
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )
    right = Alignment(horizontal="right", vertical="center")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    def _style_header(ws, row):
        for cell in ws[row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = thin_border

    # 1. Summary Sheet
    ws = wb.active
    ws.title = "Executive Summary"
    ws.append(["CAR CASTLE GOA — FLEET & FINANCIAL SUMMARY"])
    ws["A1"].font = Font(bold=True, size=16, color="20373B")
    ws.append([f"Period: {data['period_label']}"])
    ws.append([f"Generated At: {data['generated_at']}"])
    ws.append([])
    ws.append(["Financial Metric", "Value (INR)", "Notes / Details"])
    _style_header(ws, ws.max_row)

    t = data["totals"]
    summary_rows = [
        ("Total Bookings", len(data["bookings"]), "Count of reservations in this period"),
        ("Gross Customer Revenue", t["income"], "Total sales collected / billed from customers"),
        ("Owner Rent / Fleet Payables", t["owner_cost"], "Total payables owed to car owners"),
        ("Gross Operating Margin", t["margin"], "Customer Rate minus Owner Rent"),
        ("Agent Commissions", t["agent_fee"], "Total commission fees paid to agents"),
        ("Net Operating Profit", t["net_profit"], "Gross Margin minus Agent Fees"),
        ("Total Cash Collections", t["cash_income"], "Direct cash collected"),
        ("Total Online / UPI Collections", t["online_income"], "Online / Bank / UPI collected"),
        ("Security Deposits Held", t["deposit_held"], "Active security deposits held"),
        ("Security Deposits Refunded", t["deposit_refunded"], "Refunded security deposits"),
        (f"Reserve Fund Savings ({t['savings_percent']:.0f}%)", t["savings"], "Internal company capital reserve"),
    ]
    for metric, val, note in summary_rows:
        ws.append([metric, val, note])
        r = ws.max_row
        ws[f"A{r}"].font = Font(bold=True)
        ws[f"B{r}"].alignment = right
        if isinstance(val, (int, float)) and metric != "Total Bookings":
            ws[f"B{r}"].number_format = "#,##0.00"
        for col in ("A", "B", "C"):
            ws[f"{col}{r}"].border = thin_border

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 46

    # 2. Bookings Master Sheet
    ws2 = wb.create_sheet("Bookings Master")
    headers2 = [
        "Booking ID", "Customer Name", "Customer Phone", "Start Date", "Pickup Time",
        "End Date", "Drop Time", "Days", "Car Model", "Plate / Reg No", "Owner Name",
        "Payment Mode", "Deposit Amount", "Deposit Status", "Daily Cost Rate", "Daily Customer Rate",
        "Total Owner Cost", "Total Customer Rate", "Agent Fee", "Gross Margin", "Net Profit",
        "Status", "Airport Transfer", "Flight Time", "Pickup Location", "Drop Location", "Notes"
    ]
    ws2.append(headers2)
    _style_header(ws2, 1)

    for b in data["bookings"]:
        b_days = int(b.get("days") or 1)
        d_cost = float(b.get("daily_cost_rate") or (b["cost_rate"] / b_days if b_days > 0 else b["cost_rate"]))
        d_cust = float(b.get("daily_customer_rate") or (b["customer_rate"] / b_days if b_days > 0 else b["customer_rate"]))

        row_vals = [
            b.get("id", ""),
            b.get("customer_name", ""),
            b.get("customer_contact", ""),
            b.get("start_date", "")[:10],
            b.get("pickup_time", "09:00"),
            b.get("end_date", "")[:10],
            b.get("drop_time", "09:00"),
            b_days,
            b.get("car_model", "Standard Vehicle"),
            b.get("car_registration", "TBD"),
            b.get("owner_name", "—"),
            "Cash" if b.get("payment_method") == "cash" else "Online",
            float(b.get("deposit_amount") or 0),
            b.get("deposit_status", "none"),
            d_cost,
            d_cust,
            float(b["cost_rate"]),
            float(b["customer_rate"]),
            float(b.get("agent_fee", 0)),
            float(b["margin"]),
            float(b["net_profit"]),
            b.get("status", "").replace("_", " ").title(),
            b.get("transfer_type", "none"),
            b.get("flight_time", "—"),
            b.get("pickup_location", "—"),
            b.get("drop_location", "—"),
            b.get("notes", ""),
        ]
        ws2.append(row_vals)
        r = ws2.max_row
        for col_idx in range(1, len(headers2) + 1):
            cell = ws2.cell(row=r, column=col_idx)
            cell.border = thin_border
            if col_idx in (8, 13, 15, 16, 17, 18, 19, 20, 21):
                cell.alignment = right
                if col_idx != 8:
                    cell.number_format = "#,##0.00"

    for col in ws2.columns:
        col_letter = col[0].column_letter
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws2.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # 3. Owner Payables Sheet
    ws3 = wb.create_sheet("Owner Ledgers")
    ws3.append(["Owner Name", "Contact", "Total Owed (INR)", "Total Paid (INR)", "Outstanding Balance (INR)"])
    _style_header(ws3, 1)
    for o in data["owners"]:
        owed = float(o.get("total_owed") or 0)
        paid = float(o.get("total_paid") or 0)
        bal = owed - paid
        ws3.append([o.get("name", ""), o.get("contact", ""), owed, paid, bal])
        r = ws3.max_row
        for c in (3, 4, 5):
            cell = ws3.cell(row=r, column=c)
            cell.alignment = right
            cell.number_format = "#,##0.00"
            cell.border = thin_border
        for c in (1, 2):
            ws3.cell(row=r, column=c).border = thin_border
    for col in ("A", "B", "C", "D", "E"):
        ws3.column_dimensions[col].width = 24

    # 4. Agent Commissions Sheet
    ws4 = wb.create_sheet("Agent Commissions")
    ws4.append(["Agent Name", "Contact", "Total Commission Owed", "Total Commission Paid", "Balance Due"])
    _style_header(ws4, 1)
    for a in data["agents"]:
        owed = float(a.get("total_owed") or 0)
        paid = float(a.get("total_paid") or 0)
        bal = owed - paid
        ws4.append([a.get("name", ""), a.get("contact", ""), owed, paid, bal])
        r = ws4.max_row
        for c in (3, 4, 5):
            cell = ws4.cell(row=r, column=c)
            cell.alignment = right
            cell.number_format = "#,##0.00"
            cell.border = thin_border
        for c in (1, 2):
            ws4.cell(row=r, column=c).border = thin_border
    for col in ("A", "B", "C", "D", "E"):
        ws4.column_dimensions[col].width = 24

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="car-castle-goa-{data["file_label"]}.xlsx"'},
    )
